import os
import pickle

import numpy as np
import torch
from loguru import logger
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def prediction(net, dataloader, out_idx=-1, use_sign=True):
    device = next(net.parameters()).device
    codes, clses = [], []
    net.eval()
    logger.info(f'predicting({len(dataloader.dataset)})...')
    # for img, cls, _ in tqdm(dataloader):
    for x in dataloader:
        with torch.no_grad():
            out = net(x[0].to(device))
        codes.append(out if out_idx == -1 else out[out_idx])
        clses.append(x[1])
    return torch.cat(codes).sign() if use_sign else torch.cat(codes), torch.cat(clses).to(device)


def mean_average_precision(qB, rB, qL, rL, topk=-1):
    """
    Calculate mean average precision(map).

    Args:
        qB (torch.Tensor): Query data hash code.
        rB (torch.Tensor): Database data hash code.
        qL (torch.Tensor): Query data targets, one-hot
        rL (torch.Tensor): Database data targets, one-hot
        topk (float): Calculate top k data mAP.

    Returns:
        meanAP (float): Mean Average Precision.
    """
    num_query = qL.shape[0]
    if topk == -1:
        topk = rL.shape[0]
    mean_AP = 0.0
    for i in range(num_query):
        # Retrieve images from database
        retrieval = (qL[i, :] @ rL.T > 0).float()
        # Calculate hamming distance
        hamming_dist = 0.5 * (rB.shape[1] - qB[i, :] @ rB.T)
        # Arrange position according to hamming distance
        retrieval = retrieval[torch.argsort(hamming_dist)][:topk]
        # Retrieval count
        retrieval_cnt = retrieval.sum().int().item()
        # Can not retrieve images
        if retrieval_cnt == 0:
            continue
        # Generate score for every position
        score = torch.linspace(1, retrieval_cnt, retrieval_cnt).to(retrieval.device)
        # Acquire index
        index = ((retrieval == 1).nonzero(as_tuple=False).squeeze() + 1.0).float()
        mean_AP += (score / index).mean()
    mean_AP = mean_AP / num_query
    return mean_AP


def calc_hamming_dist(B1, B2):
    q = B2.shape[1]
    if len(B1.shape) < 2:
        B1 = B1.unsqueeze(0)
    distH = 0.5 * (q - B1.mm(B2.t()))
    return distH


def pr_curve(qB, rB, query_label, retrieval_label):
    num_query = qB.shape[0]
    num_bit = qB.shape[1]
    P = torch.zeros(num_query, num_bit + 1)
    R = torch.zeros(num_query, num_bit + 1)
    for i in range(num_query):
        gnd = (query_label[i].unsqueeze(0).mm(retrieval_label.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[i, :], rB)
        tmp = (hamm <= torch.arange(0, num_bit + 1).reshape(-1, 1).float().to(hamm.device)).float()
        total = tmp.sum(dim=-1)
        total = total + (total == 0).float() * 0.1
        t = gnd * tmp
        count = t.sum(dim=-1)
        p = count / total
        r = count / tsum
        P[i] = p
        R[i] = r
    mask = (P > 0).float().sum(dim=0)
    mask = mask + (mask == 0).float() * 0.1
    P = P.sum(dim=0) / mask
    R = R.sum(dim=0) / mask
    return P, R


def p_topK(qB, rB, qL, rL, K=None):
    if K is None:
        K = [1, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    num_query = qL.shape[0]
    p = [0] * len(K)
    for iter in range(num_query):
        gnd = (qL[iter].unsqueeze(0).mm(rL.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[iter, :], rB).squeeze()
        for i in range(len(K)):
            total = min(K[i], rL.shape[0])
            ind = torch.sort(hamm)[1][:total]
            gnd_ = gnd[ind]
            p[i] += gnd_.sum() / total
    p = torch.Tensor(p) / num_query
    return p


def cos(A, B=None):
    """cosine"""
    # An = normalize(A, norm='l2', axis=1)
    An = A / np.linalg.norm(A, ord=2, axis=1)[:, np.newaxis]
    if (B is None) or (B is A):
        return np.dot(An, An.T)
    # Bn = normalize(B, norm='l2', axis=1)
    Bn = B / np.linalg.norm(B, ord=2, axis=1)[:, np.newaxis]
    return np.dot(An, Bn.T)


def hamming(A, B=None):
    """A, B: [None, bit]
    elements in {-1, 1}
    """
    if B is None:
        B = A
    bit = A.shape[1]
    return (bit - A.dot(B.T)) // 2


def euclidean(A, B=None, sqrt=False):
    aTb = np.dot(A, B.T)
    if (B is None) or (B is A):
        aTa = np.diag(aTb)
        bTb = aTa
    else:
        aTa = np.diag(np.dot(A, A.T))
        bTb = np.diag(np.dot(B, B.T))
    D = aTa[:, np.newaxis] - 2.0 * aTb + bTb[np.newaxis, :]
    if sqrt:
        D = np.sqrt(D)
    return D


def NDCG(qF, rF, qL, rL, what=0, k=-1):
    """Normalized Discounted Cumulative Gain
    ref: https://github.com/kunhe/TALR/blob/master/%2Beval/NDCG.m
    """
    n_query = qF.shape[0]
    if (k < 0) or (k > rF.shape[0]):
        k = rF.shape[0]
    Rel = np.dot(qL, rL.T).astype(int)
    G = 2 ** Rel - 1
    D = np.log2(2 + np.arange(k))
    if what == 0:
        Rank = np.argsort(1 - cos(qF, rF))
    elif what == 1:
        Rank = np.argsort(hamming(qF, rF))
    elif what == 2:
        Rank = np.argsort(euclidean(qF, rF))

    _NDCG = 0
    for g, rnk in zip(G, Rank):
        dcg_best = (np.sort(g)[::-1][:k] / D).sum()
        if dcg_best > 0:
            dcg = (g[rnk[:k]] / D).sum()
            _NDCG += dcg / dcg_best
    return _NDCG / n_query


def get_precision_recall_by_Hamming_Radius(database_output, database_labels, query_output, query_labels, radius=2):
    bit_n = query_output.shape[1]
    ips = np.dot(query_output, database_output.T)
    ips = (bit_n - ips) / 2

    precX = []
    for i in range(ips.shape[0]):
        label = query_labels[i, :]
        label[label == 0] = -1
        idx = np.reshape(np.argwhere(ips[i, :] <= radius), (-1))
        all_num = len(idx)
        if all_num != 0:
            imatch = np.sum(database_labels[idx[:], :] == label, 1) > 0
            match_num = np.sum(imatch)
            precX.append(match_num / all_num)
        else:
            precX.append(0.0)
    return np.mean(np.array(precX))


def get_proj_names():
    return ['DPSH', 'DSH', 'CSQ', 'OrthoHash', 'IDHN', 'HyP2', 'CenterHashing', 'SWTH', 'SPRCH']


def prepare_excel(file_path, sheet_name):
    is_new = False
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        is_new = True
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
    return wb, ws, is_new


def write_excel_pr(file_path, proj_name, dataset, hash_bit, P, R):
    wb, ws, is_new = prepare_excel(file_path, f"{dataset}-{hash_bit}")
    if is_new:
        proj_names = get_proj_names()
        for i, x in enumerate(proj_names):
            ws.cell(row=1, column=2 * i + 1).value = x
            ws.cell(row=2, column=2 * i + 1).value = 'R'
            ws.cell(row=2, column=2 * i + 2).value = 'P'

    row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
    try:
        j = row1.index(proj_name) + 1
    except ValueError:
        j = len(row1) + 1
        ws.cell(row=1, column=j).value = proj_name
        ws.cell(row=2, column=j).value = 'R'
        ws.cell(row=2, column=j + 1).value = 'P'

    for i, x in enumerate(R):
        ws.cell(row=i + 3, column=j).value = x.item()
        ws.cell(row=i + 3, column=j + 1).value = P[i].item()

    wb.save(file_path)


def write_excel_map_ndcg(file_path, proj_name, dataset, hash_bit, v):
    """
    type: NDCG or mAP
    """
    wb, ws, is_new = prepare_excel(file_path, proj_name)
    if is_new:
        ws.cell(row=1, column=2).value = "16bits"
        ws.cell(row=1, column=3).value = "32bits"
        ws.cell(row=1, column=4).value = "64bits"
        ws.cell(row=1, column=5).value = "128bits"
        ws.cell(row=2, column=1).value = "CIFAR"
        ws.cell(row=3, column=1).value = "NUSWIDE"
        ws.cell(row=4, column=1).value = "FLICKR"
        ws.cell(row=5, column=1).value = "COCO"

    row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
    col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

    j = row1.index(f"{hash_bit}bits") + 1
    i = col1.index(dataset.upper()) + 1

    ws.cell(row=i, column=j).value = v
    wb.save(file_path)


def write_excel_hamming2(file_path, proj_name, dataset, hash_bit, v):
    """
    save P@H≤2 result to excel.
    """
    wb, ws, is_new = prepare_excel(file_path, dataset)
    if is_new:
        proj_names = get_proj_names()
        for i in range(len(proj_names)):
            ws.cell(row=1, column=i + 2).value = proj_names[i]
        ws.cell(row=2, column=1).value = 16
        ws.cell(row=3, column=1).value = 32
        ws.cell(row=4, column=1).value = 48
        ws.cell(row=5, column=1).value = 64

    row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
    col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

    j = row1.index(proj_name) + 1
    i = col1.index(hash_bit) + 1

    ws.cell(row=i, column=j).value = v
    wb.save(file_path)


def write_excel_topk(file_path, proj_name, dataset, hash_bit, rst):
    """
    save TopN-precision result to excel.
    """
    wb, ws, is_new = prepare_excel(file_path, f"{dataset}-{hash_bit}")

    j = 1
    if is_new:
        proj_names = get_proj_names()
        for i in range(len(proj_names)):
            ws.cell(row=1, column=i + 1).value = proj_names[i]
    else:
        while True:
            if ws.cell(row=1, column=j).value is None:
                ws.cell(row=1, column=j).value = proj_name
                break
            if ws.cell(row=1, column=j).value != proj_name:
                j += 1
                continue
            break

    for i in range(len(rst)):
        ws.cell(row=i + 2, column=j).value = rst[i].item()

    wb.save(file_path)


def get_run_dic(evals, hash_bit):
    rst = {}
    if "mAP" in evals and hash_bit != 48:
        rst["mAP"] = 1
    if "PR-curve" in evals and hash_bit <= 32:
        rst["PR-curve"] = 1
    if "TopN-precision" in evals and hash_bit <= 32:
        rst["TopN-precision"] = 1
    if "NDCG" in evals and hash_bit != 48:
        rst["NDCG"] = 1
    if "P@H≤2" in evals and hash_bit != 128:
        rst["P@H≤2"] = 1
    if "NDCG" in rst or "P@H≤2" in rst:
        rst["use-np"] = 1
    return rst


def get_code_and_label(main_dir, proj_name, model_name, dataset, hash_bit, calc_code_with_label):
    cache_path = f"{main_dir}/{proj_name}/output/{model_name}/{dataset}/{hash_bit}/cache.p"

    if not os.path.exists(cache_path):

        qB, qL, rB, rL = calc_code_with_label(main_dir, proj_name, model_name, dataset, hash_bit)

        save_obj = {
            'qB': qB.cpu(),
            'qL': qL.cpu(),
            'rB': rB.cpu(),
            'rL': rL.cpu(),
        }
        with open(cache_path, 'ab') as f:
            pickle.dump(save_obj, f)

    else:
        with open(cache_path, 'rb') as f:
            data = pickle.load(f)
        qB, qL, rB, rL = data['qB'].cuda(), data['qL'].cuda(), data['rB'].cuda(), data['rL'].cuda()

    return qB, qL, rB, rL


def run_evals(main_dir, proj_name, model_name, evals, datasets, hash_bits, calc_code_with_label):
    for dataset in datasets:
        logger.info(f'processing dataset: {dataset}')

        for hash_bit in hash_bits:
            logger.info(f'processing hash-bit: {hash_bit}')

            run_dic = get_run_dic(evals, hash_bit)
            if not run_dic:
                logger.info(f'no eval to run, pass')
                continue

            qB, qL, rB, rL = get_code_and_label(main_dir, proj_name, model_name, dataset, hash_bit,
                                                calc_code_with_label)

            # calc mAP
            if "mAP" in run_dic:
                topk = 5000 if dataset == 'nuswide' else -1
                map = mean_average_precision(qB, rB, qL, rL, topk)
                write_excel_map_ndcg(f"{main_dir}/eval_map.xlsx", proj_name, dataset, hash_bit, f"{map:.3f}")
                logger.info(f"[dataset:{dataset}][bits:{hash_bit}][mAP@{topk}:{map:.3f}]")

            # calc PR curve
            if "PR-curve" in run_dic:
                P, R = pr_curve(qB, rB, qL, rL)
                write_excel_pr(f"{main_dir}/eval_pr.xlsx", proj_name, dataset, hash_bit, P, R)
                logger.info(f"[dataset:{dataset}][bits:{hash_bit}][PR-curve is done]")

            # calc TopN precision
            if "TopN-precision" in run_dic:
                rst = p_topK(qB, rB, qL, rL)
                write_excel_topk(f"{main_dir}/eval_topk.xlsx", proj_name, dataset, hash_bit, rst)
                logger.info(f"[dataset:{dataset}][bits:{hash_bit}][TopN-precision is done]")

            if "use-np" in run_dic:
                qB = qB.cpu().numpy()
                rB = rB.cpu().numpy()
                qL = qL.cpu().numpy()
                rL = rL.cpu().numpy()

            # calc NDCG
            if "NDCG" in run_dic:
                ndcg = NDCG(qB, rB, qL, rL, what=1, k=1000)
                write_excel_map_ndcg(f"{main_dir}/eval_ndcg.xlsx", proj_name, dataset, hash_bit, f"{ndcg:.3f}")
                logger.info(f"[dataset:{dataset}][bits:{hash_bit}][ndcg:{ndcg:.3f}]")

            # calc Precision curves within Hamming Radius 2
            if "P@H≤2" in run_dic:
                prec = get_precision_recall_by_Hamming_Radius(rB, rL, qB, qL)
                write_excel_hamming2(f"{main_dir}/eval_hamming2.xlsx", proj_name, dataset, hash_bit, prec)
                logger.info(f"[dataset:{dataset}][bits:{hash_bit}][P@H≤2:{prec:.3f}]")
